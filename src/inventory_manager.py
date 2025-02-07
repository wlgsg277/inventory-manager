import tkinter as tk
from tkinter import ttk, filedialog
from ttkbootstrap import Style
import pandas as pd
import os
from datetime import datetime
import traceback
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class InventoryManager:
    def __init__(self):
        self.root = tk.Tk()
        self.style = Style(theme='cosmo')
        
        self.root.title("库存管理系统")
        self.root.geometry("1000x800")
        
        # 设置变量
        self.inventory_file = tk.StringVar()
        
        self.setup_ui()
        
    def setup_ui(self):
        # 主容器
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        title_label = ttk.Label(
            main_frame, 
            text="库存管理系统", 
            font=('Helvetica', 24, 'bold')
        )
        title_label.pack(pady=20)
        
        # 文件选择区域
        files_frame = ttk.LabelFrame(main_frame, text="文件选择", padding="10")
        files_frame.pack(fill=tk.X, pady=10)
        
        # 库存文件选择
        inventory_frame = ttk.Frame(files_frame)
        inventory_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(inventory_frame, text="库存文件:").pack(side=tk.LEFT)
        self.inventory_entry = ttk.Entry(
            inventory_frame, 
            textvariable=self.inventory_file,
            width=50
        )
        self.inventory_entry.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            inventory_frame, 
            text="浏览",
            style='info.TButton',
            command=self.select_inventory
        ).pack(side=tk.LEFT)
        
        # 工作表选择
        sheet_frame = ttk.Frame(files_frame)
        sheet_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(sheet_frame, text="选择工作表:").pack(side=tk.LEFT)
        self.sheet_combobox = ttk.Combobox(
            sheet_frame,
            width=20,
            state='readonly'
        )
        self.sheet_combobox.pack(side=tk.LEFT, padx=5)
        
        # 出入库文件选择
        operation_frame = ttk.Frame(files_frame)
        operation_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(operation_frame, text="出入库文件:").pack(side=tk.LEFT)
        
        # 使用Listbox显示选择的文件
        self.files_listbox = tk.Listbox(operation_frame, height=5, width=50)
        self.files_listbox.pack(side=tk.LEFT, padx=5)
        
        # 添加文件列表的滚动条
        files_scrollbar = ttk.Scrollbar(operation_frame)
        files_scrollbar.pack(side=tk.LEFT, fill=tk.Y)
        
        self.files_listbox.config(yscrollcommand=files_scrollbar.set)
        files_scrollbar.config(command=self.files_listbox.yview)
        
        # 文件操作按钮框架
        file_buttons_frame = ttk.Frame(operation_frame)
        file_buttons_frame.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            file_buttons_frame,
            text="添加文件",
            style='info.TButton',
            command=self.add_operation_files
        ).pack(pady=2)
        
        ttk.Button(
            file_buttons_frame,
            text="清除所选",
            style='danger.TButton',
            command=self.remove_selected_files
        ).pack(pady=2)
        
        ttk.Button(
            file_buttons_frame,
            text="清除全部",
            style='danger.TButton',
            command=self.clear_all_files
        ).pack(pady=2)
        
        # 更新按钮
        ttk.Button(
            main_frame, 
            text="更新库存",
            style='success.TButton',
            command=self.update_inventory,
            width=20
        ).pack(pady=20)
        
        # 日志区域
        log_frame = ttk.LabelFrame(main_frame, text="操作日志", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(log_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.log_text = tk.Text(
            log_frame, 
            height=10,
            yscrollcommand=scrollbar.set,
            wrap=tk.WORD
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.log_text.yview)
    
    def select_inventory(self):
        filename = filedialog.askopenfilename(
            title="选择库存文件",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if filename:
            self.inventory_file.set(filename)
            self.log("已选择库存文件: " + filename)
            # 读取并更新工作表列表
            try:
                excel_file = pd.ExcelFile(filename)
                self.sheet_combobox['values'] = excel_file.sheet_names
                if len(excel_file.sheet_names) > 0:
                    self.sheet_combobox.set(excel_file.sheet_names[0])
            except Exception as e:
                self.log(f"读取工作表列表时出错: {str(e)}")
    
    def add_operation_files(self):
        filenames = filedialog.askopenfilenames(
            title="选择出入库文件",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if filenames:
            for filename in filenames:
                self.files_listbox.insert(tk.END, filename)
                self.log("已添加文件: " + filename)
    
    def remove_selected_files(self):
        selection = self.files_listbox.curselection()
        for index in reversed(selection):
            self.files_listbox.delete(index)
    
    def clear_all_files(self):
        self.files_listbox.delete(0, tk.END)
    
    def log(self, message):
        self.log_text.insert(tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {message}\n")
        self.log_text.see(tk.END)
    
    def update_inventory(self):
        try:
            if not all([self.inventory_file.get(), self.sheet_combobox.get()]):
                self.log("错误: 请选择库存文件和工作表")
                return
                
            if self.files_listbox.size() == 0:
                self.log("错误: 请添加出入库文件")
                return

            progress_window = tk.Toplevel(self.root)
            progress_window.title("处理进度")
            progress_window.geometry("300x150")
            progress_window.transient(self.root)
            
            progress_label = ttk.Label(progress_window, text="正在处理...")
            progress_label.pack(pady=10)
            
            progress_bar = ttk.Progressbar(
                progress_window,
                length=200,
                mode='determinate'
            )
            progress_bar.pack(pady=10)
            
            cancel_button = ttk.Button(
                progress_window,
                text="取消",
                command=lambda: setattr(self, '_cancel_operation', True)
            )
            cancel_button.pack(pady=5)
            
            self._cancel_operation = False

            # 读取库存文件表头
            df_header = pd.read_excel(self.inventory_file.get(), sheet_name=self.sheet_combobox.get(), nrows=1)
            code_column_idx = df_header.columns.get_loc('新商品编码') + 1
            
            # 加载工作簿，保留公式
            wb = load_workbook(self.inventory_file.get(), read_only=False, data_only=False)
            ws = wb[self.sheet_combobox.get()]
            
            # 从工作表名称中获取年月信息
            sheet_name = self.sheet_combobox.get()
            try:
                sheet_date = pd.to_datetime(sheet_name.replace('年', '-').replace('月', ''))
                target_year = sheet_date.year
                target_month = sheet_date.month
                
                # 计算上个月的年月
                if target_month == 1:
                    prev_year = target_year - 1
                    prev_month = 12
                else:
                    prev_year = target_year
                    prev_month = target_month - 1
                
                prev_sheet_name = f"{prev_year}年{prev_month}月"
            except:
                self.log(f"警告: 无法从工作表名称 '{sheet_name}' 解析年月信息，将使用当前日期")
                current_date = datetime.now()
                target_year = current_date.year
                target_month = current_date.month
                prev_sheet_name = None

            # 计算30天内的日期范围
            current_date = datetime(target_year, target_month, 1)  # 当前月份的第一天
            last_processed_day = 1  # 默认为1号
            
            # 找到当前月份最后一个有出库数据的日期
            for idx, cell in enumerate(ws[1], 1):
                if cell.value and '出库' in str(cell.value):
                    try:
                        day = int(''.join(filter(str.isdigit, str(cell.value))))
                        last_processed_day = max(last_processed_day, day)
                    except:
                        continue
            
            current_date = datetime(target_year, target_month, last_processed_day)
            thirty_days_ago = current_date - pd.Timedelta(days=29)  # 获取30天前的日期
            self.log(f"计算30天出库数量：从 {thirty_days_ago.strftime('%Y年%m月%d日')} 到 {current_date.strftime('%Y年%m月%d日')}")

            # 获取上月出库数据
            prev_month_data = {}
            if prev_sheet_name and thirty_days_ago.month != target_month:
                try:
                    prev_wb = load_workbook(self.inventory_file.get(), read_only=True)
                    if prev_sheet_name in prev_wb.sheetnames:
                        prev_ws = prev_wb[prev_sheet_name]
                        
                        # 找到新商品编码列和出库数据列
                        code_col = None
                        date_cols = {}
                        
                        for idx, cell in enumerate(prev_ws[1], 1):
                            if cell.value == '新商品编码':
                                code_col = idx
                            elif cell.value and '出库' in str(cell.value):
                                try:
                                    # 从列名中提取日期
                                    day = int(''.join(filter(str.isdigit, str(cell.value))))
                                    # 只收集30天范围内的日期列
                                    check_date = datetime(prev_year, prev_month, day)
                                    if check_date >= thirty_days_ago:
                                        date_cols[idx] = day
                                except:
                                    continue
                        
                        if code_col and date_cols:
                            self.log(f"从上月收集 {len(date_cols)} 天的出库数据")
                            # 读取上月数据
                            for row in prev_ws.iter_rows(min_row=2):
                                code = str(row[code_col-1].value).strip() if row[code_col-1].value else None
                                if code:
                                    # 累计该商品所有出库数量
                                    total_out = 0
                                    for col_idx, day in date_cols.items():
                                        value = row[col_idx-1].value
                                        if value and isinstance(value, (int, float)):
                                            total_out += value
                                    
                                    if total_out > 0:
                                        prev_month_data[code] = total_out
                        
                        prev_wb.close()
                        self.log(f"已读取上月出库数据: {prev_sheet_name}")
                except Exception as e:
                    self.log(f"读取上月数据时出错: {str(e)}")

            # 建立商品编码索引
            self.log("正在建立商品编码索引...")
            code_index_map = {}
            code_column_letter = get_column_letter(code_column_idx)
            
            # 处理商品编码，确保格式一致
            for idx, row in enumerate(ws.iter_rows(min_col=code_column_idx, max_col=code_column_idx, min_row=2), 2):
                if row[0].value:
                    # 统一商品编码格式：去除空格，转为字符串
                    code = str(row[0].value).strip()
                    code_index_map[code] = idx

            total_files = self.files_listbox.size()
            progress_bar['maximum'] = total_files * 100
            
            # 预处理：收集所有更新
            all_updates = {}
            
            # 处理每个出入库文件
            for file_idx in range(total_files):
                if self._cancel_operation:
                    self.log("操作已取消")
                    break
                    
                operation_file = self.files_listbox.get(file_idx)
                file_name = os.path.basename(operation_file)
                progress_label.config(text=f"正在处理: {file_name}")
                self.log(f"\n开始处理文件: {file_name}")
                
                try:
                    # 读取文件
                    df = pd.read_excel(
                        operation_file,
                        usecols=lambda x: x in ['出库单号', '商品编码', '数量', '出库日期', '调拨日期']
                    )
                    
                    is_outbound = '出库单号' in df.columns
                    operation_type = "出库" if is_outbound else "入库"
                    date_column = '出库日期' if is_outbound else '调拨日期'
                    quantity_column = '数量'  # 都使用数量列
                    
                    if date_column not in df.columns:
                        self.log(f"警告: 文件中未找到{date_column}列")
                        continue
                    
                    # 修改日期格式识别
                    try:
                        df[date_column] = pd.to_datetime(df[date_column])
                    except:
                        self.log(f"警告: 日期格式转换失败: {date_column}")
                        continue
                    
                    # 处理商品编码格式
                    df['商品编码'] = df['商品编码'].astype(str).str.strip()
                    
                    # 数据处理
                    df_sum = df.groupby([date_column, '商品编码'])[quantity_column].sum()
                    
                    for (date, code), quantity in df_sum.items():
                        # 处理目标月份的数据
                        if date.year == target_year and date.month == target_month:
                            if code in code_index_map:
                                # 修改列名格式为"1月22日出/进库"
                                column_name = f"{date.month}月{date.day}日{'出' if is_outbound else '进'}库"
                                
                                # 获取或创建日期列缓存
                                if column_name not in all_updates:
                                    found_column = False
                                    for idx, cell in enumerate(ws[1], 1):
                                        if cell.value == column_name:
                                            all_updates[column_name] = (get_column_letter(idx), {})
                                            found_column = True
                                            break
                                    if not found_column:
                                        self.log(f"警告: 未找到列 '{column_name}'")
                                        continue

                                if column_name in all_updates:
                                    column_letter, updates_dict = all_updates[column_name]
                                    # 累加相同商品编码的数量
                                    current_value = updates_dict.get(code_index_map[code], 0)
                                    updates_dict[code_index_map[code]] = current_value + quantity
                        
                        # 处理近30天出库数量
                        if is_outbound and date >= thirty_days_ago:
                            if code in code_index_map and '近30天出库数量' in all_updates:
                                column_letter, updates_dict = all_updates['近30天出库数量']
                                current_value = updates_dict.get(code_index_map[code], 0)
                                updates_dict[code_index_map[code]] = current_value + quantity
                    
                    # 更新进度
                    progress = (file_idx * 100) + 100
                    progress_bar['value'] = min(progress, progress_bar['maximum'])
                    self.root.update_idletasks()
                    
                except Exception as e:
                    self.log(f"处理文件 {file_name} 时出错: {str(e)}")
                    continue

            if not self._cancel_operation:
                # 批量应用所有更新，保留公式和格式
                self.log("\n正在更新单元格...")
                for column_name, (column_letter, updates) in all_updates.items():
                    for row_idx, value in updates.items():
                        cell = ws[f"{column_letter}{row_idx}"]
                        cell.value = value  # 只更新值，保留公式和格式
                
                # 保存更新后的库存表
                try:
                    self.log("正在保存文件...")
                    wb.save(self.inventory_file.get())
                    self.log("\n已成功保存更新后的库存表")
                except Exception as e:
                    self.log(f"错误: 保存文件时出错 - {str(e)}")
            
            progress_window.destroy()
            
        except Exception as e:
            self.log(f"错误: {str(e)}\n")
            self.log(traceback.format_exc())
            if 'progress_window' in locals():
                progress_window.destroy()
    
    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = InventoryManager()
    app.run()
